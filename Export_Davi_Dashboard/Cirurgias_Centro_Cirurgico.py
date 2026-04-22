# Data da última modificação: 2025-08-05
# Autor: @IvanReis
# Descrição: HSF - Painel Florence - Indicadores de Cirurgias

import datetime
import io
import os
import sys

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# Adiciona o diretório raiz para importar utils_florence
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from utils_florence import apply_florence_ui, ensure_oracle_initialized

# --- Configuração da Página ---
# st.set_page_config(layout="wide", initial_sidebar_state="expanded")

# Aplica o Design System Florence
apply_florence_ui()

# --- Funções de Carregamento de Dados ---

def carregar_dados_cirurgias(data_inicial, data_final):
    """Carrega os dados de cirurgias do centro cirúrgico."""
    query_file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "querys", "query_centro_cirurgico_florence.sql")
    try:
        with open(query_file_path, "r", encoding="utf-8") as f:
            query = f.read()
    except Exception as e:
        st.error(f"Erro ao ler query_centro_cirurgico.sql: {e}")
        return pd.DataFrame()

    params = {
        'DATA_INICIAL': data_inicial,
        'DATA_FINAL': data_final
    }

    try:
        conn = st.connection("oracle_db", type="sql")
        df = conn.query(query, params=params, ttl=3600, show_spinner=False)
        df.columns = [c.upper() for c in df.columns]
        return df
    except Exception as e:
        st.error(f"Erro ao executar a query: {e}")
        return pd.DataFrame()


# --- Funções de Lógica e Exibição ---

def metric_card(label, value):
    """Encapsula uma métrica Streamlit em um card Glassmorphism."""
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.metric(label, value)
    st.markdown('</div>', unsafe_allow_html=True)


def dataframe_to_excel_bytes(df: pd.DataFrame, title: str) -> bytes:
    """Converte DataFrame para Excel com formatação premium."""
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'Cirurgias'
        df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=4)
        worksheet = writer.sheets[sheet_name]
        
        # Estilo do Título
        title_font = Font(name='Calibri', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')
        end_col_letter = get_column_letter(df.shape[1])
        worksheet.merge_cells(f'A1:{end_col_letter}1')
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        
        # (Lógica original de cabeçalhos preservada para integridade do relatório)
        current_excel_col_idx = 1
        top_header_start_col = None
        prev_top_header = None
        for col_name in df.columns:
            try: col_carater, col_porte = col_name
            except ValueError: col_carater, col_porte = col_name, ''
            if col_carater == 'DATA' and col_porte == '':
                worksheet.cell(row=4, column=current_excel_col_idx, value='DATA').font = Font(bold=True)
                worksheet.merge_cells(start_row=3, start_column=current_excel_col_idx, end_row=4, end_column=current_excel_col_idx)
                if prev_top_header is not None:
                    worksheet.merge_cells(start_row=3, start_column=top_header_start_col, end_row=3, end_column=current_excel_col_idx - 1)
                    worksheet.cell(row=3, column=top_header_start_col, value=prev_top_header).font = Font(bold=True)
                prev_top_header = None
                top_header_start_col = None
            else:
                if col_carater != prev_top_header:
                    if prev_top_header is not None:
                        worksheet.merge_cells(start_row=3, start_column=top_header_start_col, end_row=3, end_column=current_excel_col_idx - 1)
                        worksheet.cell(row=3, column=top_header_start_col, value=prev_top_header).font = Font(bold=True)
                    top_header_start_col = current_excel_col_idx
                    prev_top_header = col_carater
                worksheet.cell(row=4, column=current_excel_col_idx, value=col_porte).font = Font(bold=True)
            current_excel_col_idx += 1
        
    return output.getvalue()


def exibir_kpis_resumo(df: pd.DataFrame):
    """Exibe os KPIs com Glassmorphism."""
    st.markdown("### 📊 Resumo de Volume Cirúrgico")
    
    # Cálculos
    total_geral = df['QT_TOTAL'].sum()
    total_eletivas = df[df['GRUPO_CARATER'] == 'Eletiva']['QT_TOTAL'].sum()
    total_urgencias = df[df['GRUPO_CARATER'] == 'Urgência / Emergência']['QT_TOTAL'].sum()

    cols = st.columns(3)
    with cols[0]: metric_card("⚕️ Total Geral", int(total_geral))
    with cols[1]: metric_card("📅 Eletivas", int(total_eletivas))
    with cols[2]: metric_card("🚑 Urgências", int(total_urgencias))


def exibir_tabela_cirurgias(df):
    """Exibe a tabela pivotada estilizada."""
    df_renomeado = df.rename(columns={
        'DT_CIRURGIA': 'DATA', 'QT_PORTE_1': 'Porte 1', 'QT_PORTE_2': 'Porte 2',
        'QT_PORTE_3': 'Porte 3', 'QT_PORTE_4': 'Porte 4', 'QT_TOTAL': 'Total'
    })
    
    try:
        df_p = df_renomeado.pivot_table(index='DATA', columns='GRUPO_CARATER',
                                       values=['Porte 1', 'Porte 2', 'Porte 3', 'Porte 4', 'Total'],
                                       aggfunc='sum', fill_value=0)
    except Exception as e:
        st.error(f"Erro na estruturação: {e}")
        return

    df_p.columns = df_p.columns.swaplevel(0, 1)
    df_p.sort_index(axis=1, inplace=True)
    df_f = df_p.reset_index()
    
    new_cols = [('DATA', '')] + [col for col in df_f.columns[1:]]
    df_f.columns = pd.MultiIndex.from_tuples(new_cols)
    df_f.columns.names = ['Caráter', 'Porte']
    
    # Numérico
    num_cols = [col for col in df_f.columns if col != ('DATA', '')]
    for col in num_cols:
        df_f[col] = pd.to_numeric(df_f[col], errors='coerce').fillna(0).astype(int)
    
    df_f[('DATA', '')] = pd.to_datetime(df_f[('DATA', '')]).dt.strftime('%d/%m')
    
    # Linha Total
    total_row = df_f[num_cols].sum()
    total_row[('DATA', '')] = 'Total'
    df_f = pd.concat([df_f, pd.DataFrame(total_row).T], ignore_index=True)

    # Detalhes
    st.markdown("### 🛏️ Centro Cirúrgico")
    
    def highlight_style(row):
        # Estilo para linha de Total
        is_total = row[('DATA', '')] == 'Total'
        styles = []
        for col in df_f.columns:
            base_style = 'text-align: center;'
            if is_total:
                base_style += 'font-weight: bold; background-color: #cbd5e1; color: black;'
            elif col[0] == 'Eletiva':
                base_style += 'background-color: #e0f2fe; color: black;' # Azul claro
            elif col[0] == 'Urgência / Emergência':
                base_style += 'background-color: #ffedd5; color: black;' # Laranja claro
            
            if col[1] == 'Total':
                base_style += 'font-weight: bold;'
                
            styles.append(base_style)
        return styles

    st.dataframe(df_f.style.apply(highlight_style, axis=1), hide_index=True, use_container_width=True)
    
    # Preparar exportação
    df_exp = df_f.copy()
    df_exp.columns = ['DATA'] + [f'{c[0]} - {c[1]}' for c in df_f.columns[1:]]
    return df_exp


# --- Script Principal ---
ensure_oracle_initialized()

st.title("✂️ Monitoramento de Atividade Cirúrgica")

# --- Filtros ---
with st.container(border=True):
    f1, f2, f3 = st.columns([1, 1, 1])
    with f1: data_i = st.date_input("Início:", value=datetime.date.today())
    with f2: data_f = st.date_input("Fim:", value=datetime.date.today())
    with f3:
        buscar = st.button("🔄 Sincronizar Dados", type="primary", use_container_width=True)

# --- Exibição ---
if buscar:
    with st.spinner("Conectando ao Centro Cirúrgico..."):
        df_res = carregar_dados_cirurgias(data_i, data_f)

    if df_res is not None and not df_res.empty:
        exibir_kpis_resumo(df_res)
        st.divider()
        df_exp = exibir_tabela_cirurgias(df_res)
        
        if df_exp is not None:
            excel = dataframe_to_excel_bytes(df_exp, "Centro Cirúrgico - Hospital Santa Filomena")
            st.download_button("📥 Baixar Relatório Administrativo", excel, "cirurgias_hfs.xlsx", use_container_width=True)
    else:
        st.warning("Sem dados para o período.")
else:
    st.info("Aguardando definição do período.")