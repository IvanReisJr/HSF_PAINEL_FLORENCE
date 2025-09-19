# Data da última modificação: 2025-08-05
# Autor: @IvanReis
# Descrição: HSF - Painel Florence - Indicadores de Cirurgias
# chamada: streamlit run Home.py

import streamlit as st
import pandas as pd
import datetime
import os
import io
import oracledb
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

# --- Configuração da Página ---
st.set_page_config(layout="wide", initial_sidebar_state="expanded")

# --- Funções de Conexão com o Banco ---

_ORACLE_CLIENT_INITIALIZED = False

def initialize_oracle_client():
    """
    Inicializa o cliente Oracle apenas uma vez para evitar erros.
    """
    global _ORACLE_CLIENT_INITIALIZED
    if _ORACLE_CLIENT_INITIALIZED:
        return True

    diretorio_atual = os.path.dirname(os.path.abspath(__file__))
    diretorio_raiz_projeto = os.path.dirname(diretorio_atual)
    nome_pasta = "instantclient-basiclite-windows.x64-23.6.0.24.10\\instantclient_23_6"
    caminho_instantclient = os.path.join(diretorio_raiz_projeto, nome_pasta)

    if not os.path.exists(caminho_instantclient):
        st.error("Erro crítico: Oracle Instant Client não encontrado.")
        return False

    try:
        oracledb.init_oracle_client(lib_dir=caminho_instantclient)
        _ORACLE_CLIENT_INITIALIZED = True
        return True
    except Exception as e_init:
        if "already been initialized" not in str(e_init):
            st.error(f"Erro ao inicializar o cliente Oracle: {e_init}")
            return False
        _ORACLE_CLIENT_INITIALIZED = True
        return True

# --- Funções de Carregamento de Dados (com Cache) ---

def carregar_dados_cirurgias(data_inicial, data_final):
    """
    Carrega os dados de cirurgias do centro cirúrgico com base nos filtros de data.
    """
    if not _ORACLE_CLIENT_INITIALIZED:
        st.error("Cliente Oracle não inicializado.")
        return pd.DataFrame()

    query_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "query_centro_cirurgico.sql")
    try:
        with open(query_file_path, "r", encoding="utf-8") as f:
            query = f.read()
    except Exception as e:
        st.error(f"Erro ao ler query_centro_cirurgico.sql: {e}")
        return pd.DataFrame()

    params = {
        'DATA_INICIAL': data_inicial,
        'DATA_FINAL': data_final
    }

    try:
        conn = st.connection("oracle_db", type="sql")
        # O st.connection já trata o cache (ttl=3600)
        df = conn.query(query, params=params, ttl=3600, show_spinner=False)
        df.columns = [c.upper() for c in df.columns]
        return df
    except Exception as e:
        st.error(f"Erro ao executar a query: {e}")
        return pd.DataFrame()

# --- Funções de Lógica e Exibição ---

def dataframe_to_excel_bytes(df: pd.DataFrame, title: str) -> bytes:
    """
    Converte um DataFrame para um arquivo Excel (.xlsx) em memória,
    adicionando um título e formatando o layout.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'Cirurgias'
        # Escreve o DataFrame na planilha, começando da linha 3 para dar espaço ao título
        # Escreve apenas os dados, sem cabeçalhos ou índice, a partir da linha 5 (Excel)
        df.to_excel(writer, index=False, header=False, sheet_name=sheet_name, startrow=4)
        worksheet = writer.sheets[sheet_name]

        # Define o estilo do título
        title_font = Font(name='Calibri', size=14, bold=True)
        title_alignment = Alignment(horizontal='center', vertical='center')

        # Mescla as células para o título (da coluna A até a última coluna usada)
        end_col_letter = chr(ord('A') + df.shape[1] - 1)
        # Usando get_column_letter para maior robustez
        end_col_letter = get_column_letter(df.shape[1])
        worksheet.merge_cells(f'A1:{end_col_letter}1')

        # Escreve e formata o título
        title_cell = worksheet['A1']
        title_cell.value = title
        title_cell.font = title_font
        title_cell.alignment = title_alignment # Centraliza o título

        # --- Escreve os Cabeçalhos MultiIndex ---
        # Linha 3 (Excel) para cabeçalhos de nível superior ('Caráter')
        # Linha 4 (Excel) para cabeçalhos de nível secundário ('Porte')
        
        current_excel_col_idx = 1 # Coluna atual no Excel (1-indexado)
        top_header_start_col = None
        prev_top_header = None
        
        # AQUI É ONDE O CÓDIGO TEM UM COMPORTAMENTO DIFERENTE AGORA.
        # Ele tenta ler os nomes das colunas que foram "achatados"
        for col_name in df.columns:
            # Tenta desempacotar os nomes dos níveis se a coluna for uma tupla.
            # Se não for, assume que é uma coluna de nível único, como 'DATA'.
            try:
                col_carater, col_porte = col_name
            except ValueError:
                col_carater = col_name
                col_porte = ''
            
            # Lida com a coluna 'DATA'
            if col_carater == 'DATA' and col_porte == '':
                # Escreve 'DATA' na linha do cabeçalho secundário
                worksheet.cell(row=4, column=current_excel_col_idx, value='DATA').font = Font(bold=True)
                # Mescla A3:A4 para 'DATA'
                worksheet.merge_cells(start_row=3, start_column=current_excel_col_idx,
                                      end_row=4, end_column=current_excel_col_idx)
                
                # Se havia um cabeçalho de nível superior anterior, mescla suas células agora
                if prev_top_header is not None:
                    worksheet.merge_cells(start_row=3, start_column=top_header_start_col,
                                          end_row=3, end_column=current_excel_col_idx - 1)
                    worksheet.cell(row=3, column=top_header_start_col, value=prev_top_header).font = Font(bold=True)
                    worksheet.cell(row=3, column=top_header_start_col).alignment = Alignment(horizontal='center', vertical='center')
                
                prev_top_header = None # Reseta para o próximo cabeçalho de nível superior
                top_header_start_col = None
            else:
                # Para outras colunas, lida com a mesclagem do cabeçalho de nível superior
                if col_carater != prev_top_header:
                    # Se um novo cabeçalho de nível superior começa, mescla o anterior se existir
                    if prev_top_header is not None:
                        worksheet.merge_cells(start_row=3, start_column=top_header_start_col,
                                              end_row=3, end_column=current_excel_col_idx - 1)
                        worksheet.cell(row=3, column=top_header_start_col, value=prev_top_header).font = Font(bold=True)
                        worksheet.cell(row=3, column=top_header_start_col).alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Inicia um novo cabeçalho de nível superior
                    top_header_start_col = current_excel_col_idx
                    prev_top_header = col_carater
                
                # Escreve o cabeçalho de nível secundário
                worksheet.cell(row=4, column=current_excel_col_idx, value=col_porte).font = Font(bold=True)
            
            current_excel_col_idx += 1
        
        # Após o loop, mescla o último cabeçalho de nível superior se ele existir
        if prev_top_header is not None:
            worksheet.merge_cells(start_row=3, start_column=top_header_start_col,
                                  end_row=3, end_column=current_excel_col_idx - 1)
            worksheet.cell(row=3, column=top_header_start_col, value=prev_top_header).font = Font(bold=True)
            worksheet.cell(row=3, column=top_header_start_col).alignment = Alignment(horizontal='center', vertical='center')

        # Ajusta a largura das colunas para melhor visualização
        for col_idx, column in enumerate(df.columns, 1):
            # O nome da coluna pode ser uma tupla em um MultiIndex
            # Para calcular a largura, usamos o nome do segundo nível ou 'DATA'
            column_name = 'DATA' if isinstance(column, tuple) and column == ('DATA', '') else str(column[1]) if isinstance(column, tuple) else str(column)
            # Calcula a largura necessária
            max_length = max(
                len(column_name),
                df[column].astype(str).map(len).max()
            ) if not df.empty else len(column_name)
            # Adiciona um padding e define a largura
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    processed_data = output.getvalue()
    return processed_data

def exibir_kpis_resumo(df: pd.DataFrame):
    """
    Calcula e exibe os KPIs de resumo (Total, Eletivas, Urgências).
    """
    st.subheader("Resumo do Período")
    
    # Cálculos
    total_geral = df['QT_TOTAL'].sum()
    # Filtra pelos valores corretos ('Eletiva' e 'Urgência / Emergência') e soma.
    total_eletivas = df[df['GRUPO_CARATER'] == 'Eletiva']['QT_TOTAL'].sum()
    # Filtra por 'Urgência / Emergência' para corresponder aos dados.
    total_urgencias = df[df['GRUPO_CARATER'] == 'Urgência / Emergência']['QT_TOTAL'].sum()

    # Exibição em colunas
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("🧑‍⚕️ Total de Cirurgias", int(total_geral))
    with col2:
        st.metric("📅 Cirurgias Eletivas", int(total_eletivas))
    with col3:
        st.metric("🚑 Cirurgias de Urgência", int(total_urgencias))

def exibir_tabela_cirurgias(df):
    """
    Exibe a tabela formatada com os dados de cirurgias,
    reorganizando as colunas para o formato da imagem.
    """
    # Renomear as colunas para títulos mais amigáveis
    df_renomeado = df.rename(columns={
        'DT_CIRURGIA': 'DATA',
        'QT_PORTE_1': 'Porte 1',
        'QT_PORTE_2': 'Porte 2',
        'QT_PORTE_3': 'Porte 3',
        'QT_PORTE_4': 'Porte 4',
        'QT_TOTAL': 'Total'
    })
 
    # Criar a tabela pivotada para separar Eletiva e Urgência
    try:
        df_pivotado = df_renomeado.pivot_table(
            index='DATA',
            columns='GRUPO_CARATER',
            values=['Porte 1', 'Porte 2', 'Porte 3', 'Porte 4', 'Total'],
            aggfunc='sum',
            fill_value=0
        )
    except KeyError as e:
        st.error(f"Erro ao pivotar os dados: a coluna {e} não foi encontrada no resultado da query.")
        return
 
    if df_pivotado.empty:
        st.warning("Não foi possível gerar a visualização da tabela com os dados retornados.")
        return
 
    # Inverter os níveis do MultiIndex para ter GRUPO_CARATER como nível superior
    df_pivotado.columns = df_pivotado.columns.swaplevel(0, 1)
 
    # Ordenar as colunas para que 'Eletiva' venha antes de 'Urgência' e os portes fiquem em ordem
    df_pivotado.sort_index(axis=1, inplace=True)
 
    # Resetar o índice para que 'DATA' se torne uma coluna
    df_final = df_pivotado.reset_index()

    # Converter todas as colunas para um MultiIndex para consistência, tornando 'DATA' em ('DATA', '')
    new_cols = [('DATA', '')] + [col for col in df_final.columns[1:]]
    df_final.columns = pd.MultiIndex.from_tuples(new_cols)
 
    # Renomear os nomes dos níveis do cabeçalho para clareza
    df_final.columns.names = ['Caráter', 'Porte']
 
    # Garantir que as colunas de contagem sejam numéricas e do tipo inteiro
    # Selecionar todas as colunas exceto a de data para operações numéricas
    numeric_cols = [col for col in df_final.columns if col != ('DATA', '')]
    for col in numeric_cols:
        # Converte para número, tratando erros, preenchendo NAs com 0 e convertendo para int
        df_final[col] = pd.to_numeric(df_final[col], errors='coerce').fillna(0).astype(int)

    # Formatar a coluna 'DATA' para o formato 'dd/mm' como o último passo da manipulação de dados
    df_final[('DATA', '')] = pd.to_datetime(df_final[('DATA', '')]).dt.strftime('%d/%m')

    # --- Adicionar Linha de Total ---
    # Calcular a soma de todas as colunas numéricas
    total_row = df_final[numeric_cols].sum()
    total_row[('DATA', '')] = 'Total' # Adicionar o rótulo para a coluna de data

    # Adicionar a linha de total ao DataFrame
    total_df = pd.DataFrame(total_row).T
    df_final = pd.concat([df_final, total_df], ignore_index=True)

    # --- NOVO CÓDIGO AQUI ---
    # Achata o MultiIndex para um único nível de cabeçalho, combinando os nomes para a exportação
    df_para_exportar = df_final.copy()
    new_cols_for_export = []
    for col in df_para_exportar.columns:
        if col == ('DATA', ''):
            new_cols_for_export.append('DATA')
        else:
            new_cols_for_export.append(f'{col[0]} - {col[1]}')
    df_para_exportar.columns = new_cols_for_export
    # FIM DO NOVO CÓDIGO

    # Criação da tabela detalhada no Streamlit
    st.subheader("🛏️ Centro Cirúrgico - 9 Salas")
 
    # --- Estilização para Centralização (Método de Injeção de CSS) ---
    # Injetamos CSS diretamente na página. Esta é a forma mais robusta de garantir
    # que nossos estilos de centralização sejam aplicados, sobrepondo os padrões do Streamlit.
    st.markdown("""
        <style>
            /* A célula de cabeçalho (th) com a classe .col_heading é um container flex.
               A propriedade 'justify-content' controla o alinhamento horizontal.
               Esta regra força o conteúdo de TODOS os cabeçalhos a se centralizar. */
            .stDataFrame .col_heading {
                justify-content: center !important;
            }

            /* Esta regra garante que o conteúdo das células de dados também fique centralizado. */
            .stDataFrame td {
                text-align: center !important;
            }
        </style>
    """, unsafe_allow_html=True)

    # Aplicamos apenas o estilo da linha de total, pois o alinhamento já foi tratado pelo CSS injetado.
    def highlight_total(row):
        # Acessa a coluna de data pelo nome do MultiIndex para o display
        # Adiciona negrito, cor de fundo e cor de texto para garantir a visibilidade em todos os temas.
        style = 'font-weight: bold; background-color: #f0f2f6; color: black;'
        return [style if row[('DATA', '')] == 'Total' else '' for _ in row]
 
    # Identifica as colunas para aplicar os estilos de cor de fundo
    eletiva_cols = [col for col in df_final.columns if col[0] == 'Eletiva']
    urgencia_cols = [col for col in df_final.columns if col[0] == 'Urgência / Emergência']
    total_cols = [col for col in df_final.columns if col[1] == 'Total']

    # Aplica os estilos de forma encadeada. Adicionamos 'color: black' para garantir
    # a legibilidade no modo escuro.
    styled_df = df_final.style \
        .set_properties(subset=eletiva_cols, **{'background-color': '#e6f7ff', 'color': 'black'}) \
        .set_properties(subset=urgencia_cols, **{'background-color': '#fff0e6', 'color': 'black'}) \
        .set_properties(subset=total_cols, **{'font-weight': 'bold'}) \
        .apply(highlight_total, axis=1)
    st.dataframe(styled_df, hide_index=True, use_container_width=True)

    return df_para_exportar


# --- Script Principal ---

initialize_oracle_client()

st.title("Cirurgias")

# --- Filtros Principais ---
col_data1, col_data2, col_botao = st.columns(3)

with col_data1:
    data_inicial = st.date_input("Data inicial", value=datetime.date.today())

with col_data2:
    data_final = st.date_input("Data final", value=datetime.date.today(), min_value=data_inicial)

with col_botao:
    st.markdown("<br>", unsafe_allow_html=True)
    buscar = st.button("Buscar dados", type="primary", use_container_width=True)

# --- Lógica de Exibição ---
if buscar:
    with st.spinner("Buscando dados no banco... Por favor, aguarde."):
        df_resultado = carregar_dados_cirurgias(data_inicial, data_final)

    if df_resultado is None:
        pass
    elif not df_resultado.empty:
        # Adiciona a nova seção de KPIs de resumo
        exibir_kpis_resumo(df_resultado)
        st.divider()

        # A função agora exibe a tabela e retorna o dataframe formatado para exportação
        df_para_exportar = exibir_tabela_cirurgias(df_resultado)
        st.divider()

        # Garante que o botão de download só apareça se a tabela foi gerada com sucesso
        if df_para_exportar is not None:
            titulo_excel = "Centro Cirúrgico - 9 Salas"
            excel_bytes = dataframe_to_excel_bytes(df_para_exportar, titulo_excel)
            st.download_button(
                label="📥 Baixar dados em Excel",
                data=excel_bytes,
                file_name="Indicadores_Cirurgias.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning("Nenhum dado encontrado para os filtros selecionados.")
else:
    st.info("Selecione os filtros acima e clique em 'Buscar dados'.")